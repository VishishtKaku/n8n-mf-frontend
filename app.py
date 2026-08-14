import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import os
import requests
import bcrypt
import secrets as pysecrets
from io import BytesIO
from datetime import datetime, timezone, timedelta
from supabase import create_client
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="SBI Fund Returns Dashboard", layout="wide")

SESSION_TTL_HOURS = 24 * 7  # 7 days, sliding on each login (not on every rerun)

PERIOD_ORDER = ["1W", "1M", "3M", "6M", "YTD", "1Y", "2Y", "3Y", "5Y", "10Y", "SI"]

# Print assets: pulled from Supabase Storage bucket "logos" (public) at export
# time, not from local disk -- lets the n8n bank-logo-sync workflow keep
# bank_logos populated independently of any frontend deploy. SBI logo path is
# constant; bank logos are resolved via the bank_logos table (bank_name ->
# logo_url). Missing row/URL/failed fetch all fall through to None -- same
# silent skip as before, that side of the header just stays blank, no crash.
SBI_LOGO_STORAGE_PATH = "logos/sbi_mf_logo.png"

# SEBI-mandated mutual fund disclaimer -- edit the wording here if your
# compliance team wants different phrasing. Repeats on every printed page.
MF_DISCLAIMER_TEXT = (
    "Disclaimer: Mutual Fund investments are subject to market risks, read all "
    "scheme related documents carefully. Past performance is not indicative of future returns."
)


@st.cache_data(ttl=3600, show_spinner=False)
def get_sbi_logo_bytes():
    base_url = st.secrets["SUPABASE_URL"]
    logo_url = f"{base_url}/storage/v1/object/public/{SBI_LOGO_STORAGE_PATH}"
    try:
        resp = requests.get(logo_url, timeout=5)
        resp.raise_for_status()
        return resp.content
    except Exception:
        return None


def get_bank_logo_bytes(bank_name):
    """NOT cached at this level -- the bank_name -> logo_url table lookup must
    always see the latest row. Caching here was the earlier bug: a None result
    from before logo_url was populated got cached for up to an hour, silently
    hiding a real logo that showed up moments later. Only the actual image
    download (below) is cached, keyed on the URL itself."""
    if not bank_name:
        return None
    supabase = get_client()
    try:
        result = (
            supabase.table("bank_logos")
            .select("logo_url")
            .eq("bank_name", bank_name)
            .limit(1)
            .execute()
        )
        rows = result.data
        if not rows or not rows[0].get("logo_url"):
            return None
        return _fetch_logo_bytes_by_url(rows[0]["logo_url"])
    except Exception:
        return None


@st.cache_data(ttl=3600, show_spinner=False)
def _fetch_logo_bytes_by_url(url):
    try:
        resp = requests.get(url, timeout=5)
        resp.raise_for_status()
        return resp.content
    except Exception:
        return None


def fit_logo_dims(img_bytes, max_w=190, max_h=54):
    """Scale a logo to fit within (max_w, max_h) preserving its real aspect
    ratio -- logos come in from arbitrary sources (square, wide, tall) and
    force-stretching every one into a fixed 140x40 box distorts most of them
    into an illegible smear. Returns (width, height) in px for XLImage."""
    try:
        img = PILImage.open(BytesIO(img_bytes))
        w, h = img.size
        scale = min(max_w / w, max_h / h)
        return round(w * scale), round(h * scale)
    except Exception:
        return max_w, max_h


def crop_sbi_logo(img_bytes, top_pct=0.30891, bottom_pct=0.32079):
    """Matches the manual reference file's srcRect crop (t=30.891% b=32.079%)
    -- the source PNG is a padded square canvas, this strips the whitespace
    above/below the logo-mark so it reads as a wide banner instead of a small
    padded square. Percentages are of image height, applied top and bottom."""
    try:
        img = PILImage.open(BytesIO(img_bytes)).convert("RGBA")
        w, h = img.size
        top = round(h * top_pct)
        bottom = h - round(h * bottom_pct)
        cropped = img.crop((0, top, w, bottom))
        buf = BytesIO()
        cropped.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        return img_bytes


def multiselect_with_all(container, label, options, default_all=True, key=None, help=None):
    """Multiselect that allows any combination of options (unlike a selectbox,
    which forces exactly one choice at a time) plus an 'All' entry -- picking
    All, or picking nothing, both mean 'no filter, show everything'."""
    selected = container.multiselect(
        label, ["All"] + list(options), default=["All"] if default_all else [], key=key, help=help
    )
    if not selected or "All" in selected:
        return list(options)
    return selected


@st.cache_resource
def get_client():
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_ANON_KEY"]
    return create_client(url, key)


def get_admin_client():
    """service_role key -- bypasses RLS entirely. Used only for: account/session
    tables (app_users, app_sessions -- zero anon access, sealed even if the anon
    key leaks) and admin-gated bank-approval writes (RLS denies anon writes on
    bank_fund_approvals, only app.py's is_admin_user-gated call sites reach this
    client). Key lives in secrets.toml, server-side only -- Streamlit never
    ships it to the browser."""
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_SERVICE_ROLE_KEY"]
    return create_client(url, key)


# ============================================================
# ACCOUNTS / AUTH — app_users + app_sessions tables in Supabase.
# Session token is a random 256-bit value (pysecrets.token_hex(32)) --
# not derived from anything, so it can't be forged or decoded, only
# stolen if the URL itself leaks. Kept in st.query_params so a page
# refresh doesn't log the user out. Validated against app_sessions on
# every rerun (checks expiry, joins app_users for role).
# ============================================================

def hash_password(pw):
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw, pw_hash):
    try:
        return bcrypt.checkpw(pw.encode(), pw_hash.encode())
    except Exception:
        return False


def create_user(username, password, role="employee"):
    supabase = get_admin_client()
    existing = supabase.table("app_users").select("user_id").eq("username", username).execute()
    if existing.data:
        return None, "Username already taken."
    row = {
        "username": username,
        "password_hash": hash_password(password),
        "role": role,
    }
    resp = supabase.table("app_users").insert(row).execute()
    if not resp.data:
        return None, "Signup failed -- try again."
    return resp.data[0], None


def authenticate(username, password):
    supabase = get_admin_client()
    resp = supabase.table("app_users").select("user_id,username,password_hash,role").eq(
        "username", username
    ).execute()
    if not resp.data:
        return None, "No account with that username."
    user = resp.data[0]
    if not verify_password(password, user["password_hash"]):
        return None, "Incorrect password."
    return user, None


def create_session(user_id):
    supabase = get_admin_client()
    token = pysecrets.token_hex(32)
    expires_at = (datetime.now(timezone.utc) + timedelta(hours=SESSION_TTL_HOURS)).isoformat()
    supabase.table("app_sessions").insert(
        {"token": token, "user_id": user_id, "expires_at": expires_at}
    ).execute()
    return token


def destroy_session(token):
    if not token:
        return
    supabase = get_admin_client()
    supabase.table("app_sessions").delete().eq("token", token).execute()


def validate_session(token):
    """Returns {"user_id","username","role"} or None. Deletes the row and
    returns None if expired -- dead tokens don't linger in the table."""
    if not token:
        return None
    supabase = get_admin_client()
    resp = (
        supabase.table("app_sessions")
        .select("user_id,expires_at,app_users(username,role)")
        .eq("token", token)
        .execute()
    )
    if not resp.data:
        return None
    row = resp.data[0]
    expires_at = datetime.fromisoformat(row["expires_at"])
    if datetime.now(timezone.utc) > expires_at:
        destroy_session(token)
        return None
    user = row.get("app_users") or {}
    return {"user_id": row["user_id"], "username": user.get("username"), "role": user.get("role", "employee")}


def inject_theme_css():
    """Client-side only -- a JS class toggle on <body>, no st.rerun(). CSS
    below is scoped under body.dark-theme so it's always present in the DOM
    but inert until the class is added; toggling is instant since no
    server round-trip happens. Trade-off vs the old session_state approach:
    doesn't survive an actual browser refresh (no server state backs it),
    only a full F5 resets to light. Heavy !important use because pandas
    Styler injects its own ID-scoped CSS for the returns table, and BaseWeb
    widgets (multiselect/selectbox/input/popover) ship their own theme vars
    that a plain type-selector can't outrank without it."""
    st.markdown(
        """
        <style>
        body.dark-theme .stApp, body.dark-theme [data-testid="stAppViewContainer"],
        body.dark-theme [data-testid="stHeader"], body.dark-theme [data-testid="stMain"],
        body.dark-theme .main .block-container {
            background-color: #0E1117 !important; color: #FAFAFA !important;
        }
        body.dark-theme section[data-testid="stSidebar"] { background-color: #161A23 !important; }
        body.dark-theme section[data-testid="stSidebar"] * { color: #FAFAFA !important; }
        body.dark-theme h1, body.dark-theme h2, body.dark-theme h3, body.dark-theme h4,
        body.dark-theme h5, body.dark-theme h6, body.dark-theme p, body.dark-theme span,
        body.dark-theme label, body.dark-theme div, body.dark-theme .stMarkdown {
            color: #FAFAFA !important;
        }

        /* Returns table -- pandas Styler ships its own ID-scoped CSS, needs !important to beat it */
        body.dark-theme div[data-testid="stTable"] table,
        body.dark-theme div[data-testid="stTable"] td, body.dark-theme div[data-testid="stTable"] th {
            background-color: #161A23 !important; color: #FAFAFA !important;
            border-color: #3A3F4B !important;
        }
        body.dark-theme div[data-testid="stTable"] th { background-color: #232733 !important; }

        body.dark-theme .stDataFrame { background-color: #161A23 !important; }

        /* Inputs, selects, multiselects (BaseWeb) */
        body.dark-theme div[data-baseweb="input"], body.dark-theme div[data-baseweb="select"] > div,
        body.dark-theme div[data-baseweb="base-input"], body.dark-theme input, body.dark-theme textarea {
            background-color: #232733 !important; color: #FAFAFA !important;
            border-color: #3A3F4B !important;
        }
        body.dark-theme div[data-baseweb="tag"] { background-color: #1E4B8C !important; }
        body.dark-theme div[data-baseweb="tag"] span { color: #FAFAFA !important; }
        body.dark-theme [data-baseweb="popover"], body.dark-theme [data-baseweb="popover"] * {
            background-color: #232733 !important; color: #FAFAFA !important;
        }
        body.dark-theme ul[role="listbox"], body.dark-theme li[role="option"],
        body.dark-theme div[role="option"] {
            background-color: #232733 !important; color: #FAFAFA !important;
        }
        body.dark-theme li[role="option"]:hover, body.dark-theme li[aria-selected="true"],
        body.dark-theme div[role="option"]:hover {
            background-color: #2E3440 !important;
        }

        /* Buttons, including download buttons -- default download button
        keeps a white background from Streamlit's base theme otherwise */
        body.dark-theme .stButton button, body.dark-theme .stDownloadButton button,
        body.dark-theme [data-testid="stFormSubmitButton"] button {
            background-color: #232733 !important; color: #FAFAFA !important;
            border: 1px solid #3A3F4B !important;
        }
        body.dark-theme .stButton button:hover, body.dark-theme .stDownloadButton button:hover {
            border-color: #1E4B8C !important;
        }

        body.dark-theme .stTabs [data-baseweb="tab"] { color: #FAFAFA !important; }
        body.dark-theme .stTabs [aria-selected="true"] { color: #1E4B8C !important; }
        body.dark-theme [data-testid="stCaptionContainer"], body.dark-theme .stCaption {
            color: #B0B4BC !important;
        }
        body.dark-theme .stRadio label, body.dark-theme .stCheckbox label { color: #FAFAFA !important; }
        body.dark-theme hr { border-color: #3A3F4B !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_auth_gate():
    """Blocks the whole app until a valid session exists. Returns the
    current user dict once logged in."""
    qp_token = st.query_params.get("t")
    if "auth_user" not in st.session_state:
        st.session_state.auth_user = validate_session(qp_token) if qp_token else None

    if st.session_state.auth_user:
        return st.session_state.auth_user

    st.title("SBI Fund Returns Dashboard")
    st.subheader("Sign in")
    mode = st.radio("", ["Log in", "Sign up"], horizontal=True, key="auth_mode", label_visibility="collapsed")

    if mode == "Log in":
        with st.form("login_form"):
            username = st.text_input("Username", key="login_username")
            password = st.text_input("Password", type="password", key="login_password")
            submitted = st.form_submit_button("Log in")
        if submitted:
            if not username or not password:
                st.error("Enter both username and password.")
            else:
                user, err = authenticate(username, password)
                if err:
                    st.error(err)
                else:
                    token = create_session(user["user_id"])
                    st.session_state.auth_user = {
                        "user_id": user["user_id"], "username": user["username"], "role": user["role"]
                    }
                    st.query_params["t"] = token
                    st.rerun()
    else:
        with st.form("signup_form"):
            new_username = st.text_input("Choose a username", key="signup_username")
            new_password = st.text_input("Choose a password", type="password", key="signup_password")
            confirm_password = st.text_input("Confirm password", type="password", key="signup_confirm")
            submitted = st.form_submit_button("Sign up")
        if submitted:
            if not new_username or not new_password:
                st.error("Fill in all fields.")
            elif new_password != confirm_password:
                st.error("Passwords don't match.")
            elif len(new_password) < 8:
                st.error("Password must be at least 8 characters.")
            else:
                user, err = create_user(new_username, new_password, role="employee")
                if err:
                    st.error(err)
                else:
                    st.success("Account created. New accounts default to 'employee' access -- ask an admin to upgrade you in Supabase if you need Fund Builder access.")
                    token = create_session(user["user_id"])
                    st.session_state.auth_user = {
                        "user_id": user["user_id"], "username": user["username"], "role": user["role"]
                    }
                    st.query_params["t"] = token
                    st.rerun()

    st.stop()


@st.cache_data(ttl=300)
def load_funds_and_returns():
    supabase = get_client()
    funds_resp = (
        supabase.table("funds")
        .select("scheme_code,fund_name,amc_name,fund_type,scheme_category")
        .execute()
    )
    returns_resp = (
        supabase.table("fund_returns")
        .select(
            "scheme_code,period,abs_return,ann_return,abs_return_after,ann_return_after,"
            "sip_abs_return,sip_ann_return,sip_start_date,sip_invested,sip_latest_value"
        )
        .execute()
    )
    funds_df = pd.DataFrame(funds_resp.data)
    returns_df = pd.DataFrame(returns_resp.data)
    if funds_df.empty or returns_df.empty:
        return pd.DataFrame(), funds_df
    merged = returns_df.merge(funds_df, on="scheme_code", how="left")
    merged["plan_type"] = merged["fund_name"].apply(detect_plan_type)
    merged["option_type"] = merged["fund_name"].apply(detect_option_type)
    funds_df["plan_type"] = funds_df["fund_name"].apply(detect_plan_type)
    funds_df["option_type"] = funds_df["fund_name"].apply(detect_option_type)
    return merged, funds_df


def detect_plan_type(fund_name):
    """AMFI scheme names embed plan type as free text (e.g. 'REGULAR PLAN- GROWTH',
    'DIRECT PLAN -GROWTH') -- no separate structured field for it, so detect from
    the name itself. 'direct' checked first since some names contain both words
    in unrelated contexts; order matters little in practice but direct-first is
    the safer read."""
    if not isinstance(fund_name, str):
        return "Other/Unknown"
    name = fund_name.lower()
    if "direct" in name:
        return "Direct"
    if "regular" in name:
        return "Regular"
    return "Other/Unknown"


def detect_option_type(fund_name):
    """Same free-text situation as plan type. 'IDCW' (Income Distribution cum
    Capital Withdrawal) replaced the older 'Dividend' naming SEBI-wide in 2021,
    but older scheme name strings sometimes still say 'Dividend' -- checked as
    an IDCW synonym here rather than its own bucket, since it's the same option
    under the old name, not a distinct one."""
    if not isinstance(fund_name, str):
        return "Other/Unknown"
    name = fund_name.lower()
    if "idcw" in name or "dividend" in name:
        return "IDCW"
    if "growth" in name:
        return "Growth"
    return "Other/Unknown"


def apply_return_convention(df, convention):
    """Computes abs_display/ann_display columns from the chosen gap-navigation
    convention. 'after' falls back to the on-or-before value wherever the
    on-or-after value is null (this is always true for period 'SI' -- inception
    is the literal earliest nav_history row, there's no before/after choice for
    it -- and can also happen for a fund whose history doesn't reach far enough
    forward of a target date)."""
    df = df.copy()
    if convention == "after":
        df["abs_display"] = df["abs_return_after"].combine_first(df["abs_return"])
        df["ann_display"] = df["ann_return_after"].combine_first(df["ann_return"])
    else:
        df["abs_display"] = df["abs_return"]
        df["ann_display"] = df["ann_return"]
    return df


@st.cache_data(ttl=120)
def list_registered_banks():
    """Banks table is the source of truth for 'does this bank exist', separate
    from bank_fund_approvals -- so a bank with 0 currently-approved funds still
    shows up here. Only remove_bank() deletes from this table."""
    supabase = get_client()
    resp = supabase.table("banks").select("bank_name").order("bank_name").execute()
    return [r["bank_name"] for r in resp.data]


def register_bank(bank_name):
    supabase = get_admin_client()
    supabase.table("banks").upsert({"bank_name": bank_name}, on_conflict="bank_name").execute()


def load_bank_approvals():
    supabase = get_client()
    resp = (
        supabase.table("bank_fund_approvals")
        .select("bank_name,scheme_code,approved")
        .eq("approved", True)
        .execute()
    )
    return pd.DataFrame(resp.data)


def write_bank_approvals(bank_name, scheme_codes):
    """Full-sync: selected funds become approved=true for this bank,
    any previously-approved-but-now-unselected funds become approved=false."""
    supabase = get_admin_client()
    now = datetime.now(timezone.utc).isoformat()

    existing = (
        supabase.table("bank_fund_approvals")
        .select("scheme_code")
        .eq("bank_name", bank_name)
        .execute()
    )
    existing_codes = {r["scheme_code"] for r in existing.data}
    selected_codes = set(scheme_codes)

    to_unapprove = existing_codes - selected_codes
    rows = [
        {"bank_name": bank_name, "scheme_code": int(c), "approved": True, "updated_at": now}
        for c in selected_codes
    ] + [
        {"bank_name": bank_name, "scheme_code": int(c), "approved": False, "updated_at": now}
        for c in to_unapprove
    ]
    if not rows:
        return 0
    supabase.table("bank_fund_approvals").upsert(rows, on_conflict="bank_name,scheme_code").execute()
    return len(rows)


def add_bank_approvals(bank_name, scheme_codes):
    """Approves the given funds for a bank -- additive only, never touches
    any other existing approval for this bank. Registers the bank in the
    `banks` table too, so it persists even if later dropped to 0 funds."""
    register_bank(bank_name)
    if not scheme_codes:
        return 0
    supabase = get_admin_client()
    now = datetime.now(timezone.utc).isoformat()
    rows = [
        {"bank_name": bank_name, "scheme_code": int(c), "approved": True, "updated_at": now}
        for c in scheme_codes
    ]
    supabase.table("bank_fund_approvals").upsert(rows, on_conflict="bank_name,scheme_code").execute()
    return len(rows)


def remove_selected_bank_approvals(bank_name, scheme_codes):
    """Unapproves only the specific funds given -- every other approval for
    this bank is left untouched."""
    if not scheme_codes:
        return 0
    supabase = get_admin_client()
    now = datetime.now(timezone.utc).isoformat()
    rows = [
        {"bank_name": bank_name, "scheme_code": int(c), "approved": False, "updated_at": now}
        for c in scheme_codes
    ]
    supabase.table("bank_fund_approvals").upsert(rows, on_conflict="bank_name,scheme_code").execute()
    return len(rows)


def remove_bank(bank_name):
    """The only place a bank actually disappears -- unapproves any remaining
    funds AND deletes the `banks` registry row. Runs the delete unconditionally
    (not gated on codes being non-empty) so a bank already at 0 approved funds
    can still be removed via this path."""
    supabase = get_admin_client()
    now = datetime.now(timezone.utc).isoformat()
    existing = (
        supabase.table("bank_fund_approvals")
        .select("scheme_code")
        .eq("bank_name", bank_name)
        .execute()
    )
    codes = [r["scheme_code"] for r in existing.data]
    if codes:
        rows = [
            {"bank_name": bank_name, "scheme_code": int(c), "approved": False, "updated_at": now}
            for c in codes
        ]
        supabase.table("bank_fund_approvals").upsert(rows, on_conflict="bank_name,scheme_code").execute()
    supabase.table("banks").delete().eq("bank_name", bank_name).execute()
    return len(codes)


def fmt_pct(x):
    if pd.isna(x):
        return ""
    return f"{x * 100:.2f}%"


from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

LUMP_HEADER_LABELS = {
    "1W": "1 Week", "1M": "1 Month", "3M": "3 Month", "6M": "6 Month", "YTD": "YTD",
    "1Y": "1 Year", "2Y": "2 Years", "3Y": "3 Years", "5Y": "5 Years",
    "10Y": "10 Years", "SI": "Since Inception",
}
SIP_PERIODS = ["1Y", "2Y", "3Y", "5Y"]
SIP_LABELS = {"1Y": "1 Year", "2Y": "2 Years", "3Y": "3 Years", "5Y": "5 Years"}

TITLE_FILL = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
TITLE_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center")
THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Fixed column positions so every fund block lines up identically down the
# sheet, and every period (including 10Y/SI) always has a real column inside
# the bordered LUMP SUM grid instead of overflowing into the SIP block.
LABEL_COL = 2                                  # B — row labels
PERIOD_START_COL = 3                           # C — 1W always here, SI always at C+10
LUMP_END_COL = PERIOD_START_COL + len(PERIOD_ORDER) - 1   # M (13)
GAP_COL = LUMP_END_COL + 1                     # N — blank spacer between the two tables
SIP_START_COL = GAP_COL + 1                    # O
SIP_HEADERS = ["Period", "Rs.1000 SIP Start Date", "Invested(Rs)", "Latest(Rs)",
               "Absolute(%)", "Annualised (%)"]
SIP_END_COL = SIP_START_COL + len(SIP_HEADERS) - 1         # T (20)


@st.cache_data(ttl=300, show_spinner="Building Excel export...")
def build_formatted_workbook(subset_df, title, bank_name=None):
    """Block-per-fund layout: logo row (SBI MF logo left, bank logo right --
    repeats on every printed page via print_title_rows), title row (merged,
    bold, light-blue fill), LUMP SUM / SIP Return section-label row (styled
    the same as the title row), fund name row, then a bordered grid —
    lump-sum periods in fixed columns C..M (so 10Y/SI always land inside the
    table), a blank gap column N, then the SIP grid in O..T. Column widths
    autofit to content. Page setup targets hardcopy printing: landscape,
    fit-to-width, mutual fund disclaimer in the footer on every page."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Approved Funds"[:31]

    col_max_len = {}

    # Excel export stays black regardless of sign -- green/red is only for
    # the in-app st.table display (color_returns(), further down). colorize
    # flag on set_cell() left in place (still marks which cells are actual
    # return values vs labels/dates) but both branches now render identically.
    BLACK_FONT = Font(color="000000")
    GREEN_FONT = BLACK_FONT
    RED_FONT = BLACK_FONT

    def set_cell(row, col, value, blank_ok=False, colorize=False):
        """blank_ok=True for cells where None is a legitimate label placeholder
        (e.g. an unused row entirely) rather than a missing data point -- those
        stay truly empty. Everything else shows '-' instead of blank so it's
        clear the cell was considered, not skipped. colorize=True paints
        positive return values green, negative red -- only used on actual
        Absolute(%)/Annualised(%) return cells, not dates/amounts/labels."""
        display = value if (value is not None or blank_ok) else "-"
        cell = ws.cell(row=row, column=col, value=display)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if colorize and isinstance(display, (int, float)) and not isinstance(display, bool):
            if display > 0:
                cell.font = GREEN_FONT
            elif display < 0:
                cell.font = RED_FONT
        if display is not None:
            col_max_len[col] = max(col_max_len.get(col, 0), len(str(display)))
        return cell

    LOGO_ROW = 1
    TITLE_ROW = 2
    ws.row_dimensions[LOGO_ROW].height = 120  # matches manual reference file

    sbi_logo_bytes = get_sbi_logo_bytes()
    if sbi_logo_bytes:
        sbi_logo_bytes = crop_sbi_logo(sbi_logo_bytes)  # strip padding, matches manual crop
        img = XLImage(BytesIO(sbi_logo_bytes))
        img.width, img.height = fit_logo_dims(sbi_logo_bytes, max_w=316, max_h=117)
        img.anchor = f"{get_column_letter(LABEL_COL)}{LOGO_ROW}"
        ws.add_image(img)

    bank_logo_bytes = get_bank_logo_bytes(bank_name) if bank_name else None
    if bank_logo_bytes:
        img2 = XLImage(BytesIO(bank_logo_bytes))
        img2.width, img2.height = fit_logo_dims(bank_logo_bytes, max_w=132, max_h=132)
        img2.anchor = f"{get_column_letter(SIP_END_COL)}{LOGO_ROW}"
        ws.add_image(img2)

    ws.merge_cells(start_row=TITLE_ROW, start_column=LABEL_COL, end_row=TITLE_ROW, end_column=SIP_END_COL)
    title_cell = ws.cell(row=TITLE_ROW, column=LABEL_COL, value=title)
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = CENTER
    col_max_len[LABEL_COL] = max(col_max_len.get(LABEL_COL, 0), 20)

    r = TITLE_ROW + 2  # blank spacer row before the first fund block
    scheme_codes = subset_df["scheme_code"].unique().tolist()
    for scheme_code in scheme_codes:
        fund_rows = subset_df[subset_df["scheme_code"] == scheme_code]
        fund_name = fund_rows["fund_name"].iloc[0]
        block_top = r

        # LUMP SUM / SIP Return section row — heading style, same as title bar
        ws.merge_cells(start_row=r, start_column=LABEL_COL, end_row=r, end_column=LUMP_END_COL)
        ws.merge_cells(start_row=r, start_column=SIP_START_COL, end_row=r, end_column=SIP_END_COL)
        lc = ws.cell(row=r, column=LABEL_COL, value="LUMP SUM")
        lc.alignment = CENTER; lc.font = TITLE_FONT; lc.fill = TITLE_FILL
        sc = ws.cell(row=r, column=SIP_START_COL, value="SIP Return")
        sc.alignment = CENTER; sc.font = TITLE_FONT; sc.fill = TITLE_FILL
        r += 1

        # Fund name row — under LUMP SUM, and again under SIP Return
        ws.merge_cells(start_row=r, start_column=PERIOD_START_COL, end_row=r, end_column=LUMP_END_COL)
        fn = ws.cell(row=r, column=PERIOD_START_COL, value=fund_name)
        fn.alignment = CENTER
        ws.merge_cells(start_row=r, start_column=SIP_START_COL, end_row=r, end_column=SIP_END_COL)
        fn_sip = ws.cell(row=r, column=SIP_START_COL, value=fund_name)
        fn_sip.alignment = CENTER
        r += 1

        header_row = r
        set_cell(r, LABEL_COL, "Period")
        for ci, p in enumerate(PERIOD_ORDER):
            set_cell(r, PERIOD_START_COL + ci, LUMP_HEADER_LABELS.get(p, p))
        for ci, h in enumerate(SIP_HEADERS):
            set_cell(r, SIP_START_COL + ci, h)
        r += 1

        data_start_row = r

        set_cell(r, LABEL_COL, "Absolute(%)")
        for ci, p in enumerate(PERIOD_ORDER):
            match = fund_rows[fund_rows["period"] == p]
            val = match["abs_display"].iloc[0] if not match.empty else None
            set_cell(r, PERIOD_START_COL + ci, None if val is None or pd.isna(val) else round(val * 100, 2), colorize=True)
        r += 1

        set_cell(r, LABEL_COL, "Annualised(%)")
        for ci, p in enumerate(PERIOD_ORDER):
            match = fund_rows[fund_rows["period"] == p]
            val = match["ann_display"].iloc[0] if not match.empty else None
            set_cell(r, PERIOD_START_COL + ci, None if val is None or pd.isna(val) else round(val * 100, 2), colorize=True)
        r += 1

        set_cell(r, LABEL_COL, "Category Average(%)")
        for ci in range(len(PERIOD_ORDER)):
            set_cell(r, PERIOD_START_COL + ci, None)  # no source for this yet -- shows "-"
        r += 1

        data_end_row = r - 1  # 3 lump-sum data rows: abs, ann, category avg

        def write_sip_row(row_num, sp):
            sip_row = fund_rows[fund_rows["period"] == sp]
            rec = sip_row.iloc[0] if not sip_row.empty else None
            start_date = rec["sip_start_date"] if rec is not None else None
            invested = rec["sip_invested"] if rec is not None else None
            latest_val = rec["sip_latest_value"] if rec is not None else None
            sip_abs = rec["sip_abs_return"] if rec is not None else None
            sip_ann = rec["sip_ann_return"] if rec is not None else None
            set_cell(row_num, SIP_START_COL, SIP_LABELS[sp])
            set_cell(row_num, SIP_START_COL + 1, None if pd.isna(start_date) else start_date)
            set_cell(row_num, SIP_START_COL + 2, None if pd.isna(invested) else round(invested, 2))
            set_cell(row_num, SIP_START_COL + 3, None if pd.isna(latest_val) else round(latest_val, 2))
            set_cell(row_num, SIP_START_COL + 4, None if pd.isna(sip_abs) else round(sip_abs * 100, 2), colorize=True)
            set_cell(row_num, SIP_START_COL + 5, None if pd.isna(sip_ann) else round(sip_ann * 100, 2), colorize=True)

        for si, sp in enumerate(SIP_PERIODS):
            row_num = data_start_row + si
            if row_num > data_end_row:
                break  # only 3 lump-sum rows available to align against; extra SIP periods dropped
            write_sip_row(row_num, sp)

        # 4th SIP period (5Y) has no lump-sum row to sit beside — reference file
        # has 4 SIP rows against only 3 named lump-sum rows plus header, so it
        # actually spans one row further down than the lump-sum block; match that.
        if len(SIP_PERIODS) > (data_end_row - data_start_row + 1):
            extra_row = data_end_row + 1
            sp = SIP_PERIODS[data_end_row - data_start_row + 1]
            write_sip_row(extra_row, sp)
            block_bottom = extra_row
        else:
            block_bottom = data_end_row

        # thin border grid over every cell in this fund's block, including
        # blank cells (e.g. Category Average row, unfilled SIP cells) — but
        # NOT the GAP_COL, which stays a plain blank spacer between the tables
        for row_idx in range(block_top, block_bottom + 1):
            for col_idx in list(range(LABEL_COL, LUMP_END_COL + 1)) + list(range(SIP_START_COL, SIP_END_COL + 1)):
                ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER

        r = block_bottom + 2  # blank separator row before next fund

    # Autofit: width = longest actual value in that column, not a flat default
    for col, length in col_max_len.items():
        ws.column_dimensions[get_column_letter(col)].width = max(9, min(30, length + 3))
    ws.column_dimensions[get_column_letter(GAP_COL)].width = 3

    # Print setup for hardcopy: logo+title rows repeat at the top of every
    # printed page (Excel's print titles), mutual fund disclaimer in the
    # footer -- footers repeat on every printed page automatically, regardless
    # of how many pages the export spans. Fit to page WIDTH only (fitToHeight=0
    # means unlimited pages tall) -- all columns land on one page across, rows
    # spill down onto as many pages as needed. HeaderFooter.scaleWithDoc=False
    # keeps the footer text a fixed readable size independent of the body's
    # width-scaling, so this doesn't reintroduce the earlier tiny-footer bug.
    ws.print_title_rows = f"{LOGO_ROW}:{TITLE_ROW}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.HeaderFooter.scaleWithDoc = False  # keep footer text size fixed, independent of body scale
    ws.oddFooter.center.text = MF_DISCLAIMER_TEXT
    ws.oddFooter.center.size = 10

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_fund_block(fund_rows, scheme_code):
    fund_name = fund_rows["fund_name"].iloc[0]
    amc = fund_rows["amc_name"].iloc[0]
    ftype = fund_rows["fund_type"].iloc[0]
    category = fund_rows["scheme_category"].iloc[0] if pd.notna(fund_rows["scheme_category"].iloc[0]) else "—"

    st.markdown(f"#### {fund_name}")
    st.caption(f"{amc} • {ftype} • {category}")

    periods_present = [p for p in PERIOD_ORDER if p in fund_rows["period"].values]
    by_period = fund_rows.set_index("period")
    abs_row, ann_row = {}, {}
    for p in periods_present:
        r = by_period.loc[p]
        abs_row[p] = r["abs_display"]
        ann_row[p] = r["ann_display"]
    blank_row = {p: None for p in periods_present}

    table = pd.DataFrame(
        [abs_row, ann_row, blank_row, blank_row],
        index=["Absolute (%)", "Annualised (%)", "Category Average (%) — not available yet",
               "Scheme Benchmark (%) — not available yet"],
    )
    table = table[periods_present]

    def color_returns(val):
        if pd.isna(val):
            return ""
        return "color: #1E7B34; font-weight: 600" if val > 0 else (
            "color: #C00000; font-weight: 600" if val < 0 else ""
        )

    styled = table.style.format(lambda v: fmt_pct(v) if pd.notna(v) else "—").map(color_returns)
    st.table(styled)
    st.markdown("")


current_user = render_auth_gate()
inject_theme_css()

top1, top2, top3, top4 = st.columns([3, 1, 1, 1])
with top1:
    st.title("SBI Fund Returns Dashboard")
with top2:
    if st.button("Refresh data"):
        st.cache_data.clear()
with top3:
    components.html(
        """
        <style>
        html, body { margin: 0; padding: 0; }
        </style>
        <button id="theme-toggle-btn" style="
            width: 100%; box-sizing: border-box; padding: 0.375rem 1rem;
            border-radius: 0.5rem; border: 1px solid rgba(49, 51, 63, 0.2);
            background-color: #FFFFFF; color: #31333F; font-size: 1rem;
            line-height: 1.6; cursor: pointer; font-family: inherit;
        ">Dark mode</button>
        <script>
        // components.html runs in its own iframe -- window.parent reaches the
        // main Streamlit document (same-origin, allowed). st.markdown's
        // onclick attribute gets stripped by Streamlit's HTML sanitizer even
        // with unsafe_allow_html=True, this is the actual workaround.
        const btn = document.getElementById('theme-toggle-btn');
        btn.addEventListener('click', function () {
            const body = window.parent.document.body;
            const isDark = body.classList.toggle('dark-theme');
            btn.textContent = isDark ? 'Light mode' : 'Dark mode';
            btn.style.backgroundColor = isDark ? '#232733' : '#FFFFFF';
            btn.style.color = isDark ? '#FAFAFA' : '#31333F';
            btn.style.borderColor = isDark ? '#3A3F4B' : 'rgba(49, 51, 63, 0.2)';
        });
        </script>
        """,
        height=39,
    )
with top4:
    if st.button(f"Log out ({current_user['username']})", key="logout_btn"):
        destroy_session(st.query_params.get("t"))
        st.session_state.auth_user = None
        st.query_params.clear()
        st.rerun()

with st.spinner("Loading fund data..."):
    data, funds_only_df = load_funds_and_returns()

if data.empty:
    st.warning(
        "No data found. Check that the n8n pipeline has run and populated the "
        "funds/fund_returns tables in Supabase."
    )
    st.stop()

st.sidebar.header("Return calculation convention")
convention_choice = st.sidebar.radio(
    "For periods that land on a weekend/holiday, which nearby trading day's NAV "
    "should be used?",
    ["Nearest prior trading day", "Nearest next trading day"],
    index=0,
    key="convention_choice",
    help=(
        "Both are legitimate, commonly used conventions -- they only produce "
        "materially different numbers when a data gap sits next to a sharp NAV "
        "move, mostly visible on short periods (1M/3M). 'Since Inception' is "
        "unaffected either way -- it always uses the fund's actual earliest "
        "recorded NAV, there's no gap to navigate around for that one."
    ),
)
convention = "after" if convention_choice == "Nearest next trading day" else "before"
data = apply_return_convention(data, convention)

is_admin_user = current_user["role"] == "admin"

if is_admin_user:
    tab_all, tab_bank, tab_builder = st.tabs(
        ["All Funds", "Bank Approved Funds", "Fund Builder (Admin)"]
    )
else:
    tab_all, tab_bank = st.tabs(["All Funds", "Bank Approved Funds"])

# ============================================================
# TAB 1 — All funds (unchanged from before)
# ============================================================
with tab_all:
    st.sidebar.header("Filters — All Funds")
    amc_options = sorted(data["amc_name"].dropna().unique().tolist())
    selected_amc = multiselect_with_all(st.sidebar, "AMC", amc_options, key="amc_all")

    fund_type_options = sorted(data["fund_type"].dropna().unique().tolist())
    selected_fund_type = multiselect_with_all(st.sidebar, "Fund Type", fund_type_options, key="ftype_all")

    selected_plan_type = multiselect_with_all(
        st.sidebar, "Plan Type", ["Regular", "Direct", "Other/Unknown"], key="plan_all",
        help="Detected from the fund name text (AMFI doesn't expose this as a separate field)."
    )
    selected_option_type = multiselect_with_all(
        st.sidebar, "Option", ["Growth", "IDCW", "Other/Unknown"], key="option_all",
        help="Detected from the fund name text. 'IDCW' includes older 'Dividend'-named schemes."
    )
    fund_name_search = st.sidebar.text_input("Search fund name", key="search_all")
    max_funds = st.sidebar.number_input(
        "Max funds to display", min_value=1, max_value=200, value=15, step=5, key="max_all",
        help="529 funds is too many to render at once — narrow with filters/search or raise this."
    )

    filtered = data[
        data["amc_name"].isin(selected_amc)
        & data["fund_type"].isin(selected_fund_type)
        & data["plan_type"].isin(selected_plan_type)
        & data["option_type"].isin(selected_option_type)
    ]
    if fund_name_search:
        filtered = filtered[filtered["fund_name"].str.contains(fund_name_search, case=False, na=False)]

    all_matching_funds = filtered["scheme_code"].unique().tolist()
    st.subheader(f"{len(all_matching_funds)} funds match filters — showing first {min(max_funds, len(all_matching_funds))}")
    st.caption(f"Showing returns using: **{convention_choice}** convention. Change this in the sidebar.")

    csv_export = filtered[[
        "fund_name", "amc_name", "fund_type", "plan_type", "option_type", "scheme_category", "period",
        "abs_display", "ann_display", "abs_return", "ann_return",
        "abs_return_after", "ann_return_after", "sip_abs_return", "sip_ann_return",
    ]].rename(columns={
        "abs_display": f"abs_return_selected ({convention_choice})",
        "ann_display": f"ann_return_selected ({convention_choice})",
    }).copy()
    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            label=f"Download CSV ({len(all_matching_funds)} funds, all periods)",
            data=csv_export.to_csv(index=False).encode("utf-8"),
            file_name="sbi_fund_returns_all.csv",
            mime="text/csv",
            key="dl_all",
        )
    with dl2:
        xlsx_buf = build_formatted_workbook(filtered, "All Funds")
        st.download_button(
            label="Download Excel (formatted)",
            data=xlsx_buf,
            file_name="sbi_fund_returns_all.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_all_xlsx",
        )
    st.divider()

    for scheme_code in all_matching_funds[:max_funds]:
        render_fund_block(filtered[filtered["scheme_code"] == scheme_code], scheme_code)

    st.caption(
        "Category Average and Scheme Benchmark rows are placeholders — that data isn't "
        "computed/sourced yet on the backend."
    )

# ============================================================
# TAB 2 — Bank Approved Funds
# ============================================================
with tab_bank:
    approvals = load_bank_approvals()

    if approvals.empty:
        st.info("No bank approvals recorded yet. Add some in the Fund Builder (Admin) tab.")
    else:
        bank_options = sorted(approvals["bank_name"].unique().tolist())
        selected_bank = st.selectbox("Select bank", bank_options, key="bank_select")

        approved_codes = approvals[approvals["bank_name"] == selected_bank]["scheme_code"].tolist()
        bank_data_all_plans = data[data["scheme_code"].isin(approved_codes)]

        selected_plan_type_bank = multiselect_with_all(
            st, "Plan Type", ["Regular", "Direct", "Other/Unknown"], key="plan_bank",
            help="Detected from the fund name text (AMFI doesn't expose this as a separate field)."
        )
        selected_option_type_bank = multiselect_with_all(
            st, "Option", ["Growth", "IDCW", "Other/Unknown"], key="option_bank",
            help="Detected from the fund name text. 'IDCW' includes older 'Dividend'-named schemes."
        )
        bank_data = bank_data_all_plans[
            bank_data_all_plans["plan_type"].isin(selected_plan_type_bank)
            & bank_data_all_plans["option_type"].isin(selected_option_type_bank)
        ]

        matching_codes = bank_data["scheme_code"].unique().tolist()
        st.subheader(f"{selected_bank} — {len(matching_codes)} approved funds")
        st.caption(f"Showing returns using: **{convention_choice}** convention. Change this in the sidebar.")

        csv_bank = bank_data[[
            "fund_name", "amc_name", "fund_type", "plan_type", "option_type", "scheme_category", "period",
            "abs_display", "ann_display", "abs_return", "ann_return",
            "abs_return_after", "ann_return_after", "sip_abs_return", "sip_ann_return",
        ]].rename(columns={
            "abs_display": f"abs_return_selected ({convention_choice})",
            "ann_display": f"ann_return_selected ({convention_choice})",
        }).copy()
        dlb1, dlb2 = st.columns(2)
        with dlb1:
            st.download_button(
                label=f"Download CSV ({selected_bank})",
                data=csv_bank.to_csv(index=False).encode("utf-8"),
                file_name=f"sbi_fund_returns_{selected_bank.replace(' ', '_')}.csv",
                mime="text/csv",
                key="dl_bank",
            )
        with dlb2:
            xlsx_buf_bank = build_formatted_workbook(
                bank_data, f"{selected_bank} Approved SBI MUTUAL FUNDS", bank_name=selected_bank
            )
            st.download_button(
                label="Download Excel (formatted)",
                data=xlsx_buf_bank,
                file_name=f"sbi_fund_returns_{selected_bank.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_bank_xlsx",
            )
        st.divider()

        if not matching_codes:
            st.info("No funds in the returns data match this bank's approved list yet.")
        for scheme_code in matching_codes:
            render_fund_block(bank_data[bank_data["scheme_code"] == scheme_code], scheme_code)

# ============================================================
# TAB 3 — Fund Builder (Admin) — gated by account role, not a shared password
# ============================================================
if is_admin_user:
    with tab_builder:
        st.success(f"Logged in as admin: {current_user['username']}")

        st.markdown("### Manage Bank Approved Funds")

        existing_banks = list_registered_banks()
        bank_choice_mode = st.radio(
            "Bank", ["Choose existing", "Add new", "Remove bank"], horizontal=True, key="bank_mode"
        )

        if bank_choice_mode == "Remove bank":
            if not existing_banks:
                st.caption("No banks to remove yet.")
            else:
                remove_bank_choice = st.selectbox(
                    "Bank to remove", existing_banks, key="bank_remove_select"
                )
                st.caption(
                    f"This unapproves all funds currently listed under **{remove_bank_choice}** "
                    "— the bank disappears from Bank Approved Funds and Fund Builder dropdowns. "
                    "Type the bank name below to confirm."
                )
                confirm_text = st.text_input(
                    "Type the bank name to confirm removal", key="bank_remove_confirm"
                )
                if st.button("Remove bank", key="remove_bank_btn", type="primary"):
                    if confirm_text != remove_bank_choice:
                        st.error("Typed name doesn't match. Nothing removed.")
                    else:
                        n = remove_bank(remove_bank_choice)
                        st.cache_data.clear()
                        st.success(f"Removed {remove_bank_choice} — {n} fund approvals cleared.")
                        st.rerun()

        else:
            if bank_choice_mode == "Choose existing" and existing_banks:
                bank_name = st.selectbox("Existing bank", existing_banks, key="bank_existing")
            else:
                bank_name = st.text_input("New bank name", key="bank_new")

            all_fund_options = funds_only_df.dropna(subset=["fund_name"]).sort_values("fund_name")
            fund_label_to_code = dict(zip(all_fund_options["fund_name"], all_fund_options["scheme_code"]))
            code_to_label = {c: n for n, c in fund_label_to_code.items()}

            current_approvals_df = load_bank_approvals()
            approved_codes = set()
            if bank_name and not current_approvals_df.empty:
                approved_codes = set(
                    current_approvals_df[current_approvals_df["bank_name"] == bank_name]["scheme_code"]
                )
            approved_labels = sorted(code_to_label[c] for c in approved_codes if c in code_to_label)

            st.markdown("#### Add funds")
            addable_df = all_fund_options[~all_fund_options["scheme_code"].isin(approved_codes)]

            fc1, fc2 = st.columns(2)
            with fc1:
                plan_choice = multiselect_with_all(
                    st, "Plan Type", ["Regular", "Direct", "Other/Unknown"], key="plan_add",
                    help="Choose Growth/Direct/etc first, then narrow further by typing the fund name below.",
                )
            with fc2:
                option_choice = multiselect_with_all(
                    st, "Option", ["Growth", "IDCW", "Other/Unknown"], key="option_add",
                )
            addable_df = addable_df[
                addable_df["plan_type"].isin(plan_choice) & addable_df["option_type"].isin(option_choice)
            ]

            to_add_labels = st.multiselect(
                "Select funds to add", options=sorted(addable_df["fund_name"].tolist()),
                default=[], key="fund_add_multiselect",
                help="Type to search within this list -- narrowed by Plan Type/Option above.",
            )
            if st.button("Add selected funds", key="add_funds_btn"):
                if not bank_name:
                    st.error("Enter a bank name first.")
                elif not to_add_labels:
                    st.error("Select at least one fund to add.")
                else:
                    codes = [fund_label_to_code[lbl] for lbl in to_add_labels]
                    n = add_bank_approvals(bank_name, codes)
                    st.cache_data.clear()
                    st.success(f"Added {n} fund(s) to {bank_name}'s approval list.")
                    st.rerun()

            st.markdown("#### Currently approved funds")
            if not approved_labels:
                st.caption("No funds approved for this bank yet.")
            else:
                st.dataframe(
                    pd.DataFrame({"Approved fund": approved_labels}),
                    use_container_width=True, hide_index=True,
                    height=min(300, 40 + 35 * len(approved_labels)),
                )

                st.markdown("#### Remove funds")
                to_remove_labels = st.multiselect(
                    "Select approved funds to disapprove", options=approved_labels, default=[],
                    key="fund_remove_multiselect",
                )
                if st.button("Remove selected funds", key="remove_funds_btn", type="primary"):
                    if not to_remove_labels:
                        st.error("Select at least one fund to remove.")
                    else:
                        codes = [fund_label_to_code[lbl] for lbl in to_remove_labels]
                        n = remove_selected_bank_approvals(bank_name, codes)
                        st.cache_data.clear()
                        st.success(f"Removed {n} fund(s) from {bank_name}'s approval list.")
                        st.rerun()

        st.divider()
        st.markdown("### Current approvals (all banks)")
        current_all = load_bank_approvals()
        if current_all.empty:
            st.info("No approvals recorded yet.")
        else:
            display_current = current_all.merge(
                funds_only_df[["scheme_code", "fund_name"]], on="scheme_code", how="left"
            )[["bank_name", "fund_name"]].sort_values(["bank_name", "fund_name"])
            st.dataframe(display_current, use_container_width=True, hide_index=True)
            st.download_button(
                label="Download CSV (all bank approvals)",
                data=display_current.to_csv(index=False).encode("utf-8"),
                file_name="bank_approvals_all.csv",
                mime="text/csv",
                key="dl_builder",
            )
